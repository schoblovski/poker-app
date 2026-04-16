"""
DTKS Poker – Datenmigration
============================
Liest die Excel-Datei und importiert ALLE Daten in Supabase.
Bestehende Daten werden VOR dem Import automatisch gelöscht.

Fotos (Profilbilder + Beweisfotos):
  - Relativer Pfad in Excel → Bild wird aus FOTO_ORDNER geladen,
    komprimiert (JPEG) und als Base64 in Supabase gespeichert.
  - Bereits http-URLs oder data:-Strings werden unverändert übernommen.
  - Fehlende Dateien erzeugen eine Warnung, Migration läuft weiter.

Installation (einmalig):
  pip install openpyxl supabase pillow

Ausführen:
  python migrate_poker.py
"""

import sys
import os
import io
import base64
from openpyxl import load_workbook
from supabase import create_client
from collections import defaultdict

try:
    from PIL import Image
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ── KONFIGURATION ─────────────────────────────────────
SUPABASE_URL = 'https://bcvyhlzjpfezokvcjksn.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJjdnlobHpqcGZlem9rdmNqa3NuIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTc0MDM0OSwiZXhwIjoyMDkxMzE2MzQ5fQ.m-wDN0Ahkpkp0W-WPpYtuDuZcqAXZuT0Z-UNIoiXqSk'
EXCEL_FILE   = 'poker tracker v3.xlsx'

# Pfad zum heruntergeladenen Foto-Ordner aus Google Drive.
# Alle relativen Pfade in der Excel werden relativ zu diesem Ordner aufgelöst.
# Beispiel: FOTO_ORDNER = 'fotos'  → Datei 'spieler/chris.jpg' → 'fotos/spieler/chris.jpg'
FOTO_ORDNER  = 'fotos'

# Bildgrössen (entsprechen den App-Einstellungen)
PROFIL_MAX_PX  = 512    # Profilbilder: max 512×512px (wie im App-Upload)
BEWEIS_MAX_PX  = 1024   # Beweisfotos:  max 1024×1024px (wie im App-Upload)
JPEG_QUALITY   = 82     # JPEG-Qualität (0–100)
# ─────────────────────────────────────────────────────


def foto_zu_base64(relativer_pfad, max_px):
    """
    Lädt ein Bild, komprimiert es und gibt einen Base64-Data-URI-String zurück.
    - Relative Pfade werden aus FOTO_ORDNER aufgelöst.
    - http-URLs und data:-Strings werden unverändert zurückgegeben.
    - Fehlende Dateien: None + Warnung.
    """
    if not relativer_pfad:
        return None

    s = str(relativer_pfad).strip()

    # Bereits URL oder Base64 → unverändert übernehmen
    if s.startswith('http') or s.startswith('data:'):
        return s

    # Datei suchen: erst relativ zu FOTO_ORDNER, dann direkt
    kandidaten = [
        os.path.join(FOTO_ORDNER, s),
        s,
        os.path.join(FOTO_ORDNER, os.path.basename(s)),
    ]
    pfad = next((p for p in kandidaten if os.path.isfile(p)), None)

    if not pfad:
        print(f"    ⚠️  Foto nicht gefunden: {s}")
        return None

    if not PIL_OK:
        print(f"    ⚠️  Pillow nicht installiert – Foto wird als Pfad gespeichert: {s}")
        return s

    try:
        img = Image.open(pfad).convert('RGB')
        img.thumbnail((max_px, max_px), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=JPEG_QUALITY, optimize=True)
        b64 = base64.b64encode(buf.getvalue()).decode()
        kb = len(buf.getvalue()) // 1024
        print(f"    🖼  {os.path.basename(pfad)} → {img.width}×{img.height}px, {kb} KB")
        return f'data:image/jpeg;base64,{b64}'
    except Exception as e:
        print(f"    ⚠️  Fehler beim Verarbeiten von {s}: {e}")
        return None


def main():
    print("🃏 DTKS Poker – Datenmigration")
    print("=" * 50)

    # Pillow-Warnung
    if not PIL_OK:
        print("\n⚠️  Pillow nicht installiert – Fotos werden NICHT konvertiert!")
        print("   Installieren mit:  pip install pillow\n")

    # Foto-Ordner prüfen
    foto_ordner_ok = os.path.isdir(FOTO_ORDNER)
    if not foto_ordner_ok:
        print(f"\n⚠️  Foto-Ordner '{FOTO_ORDNER}' nicht gefunden.")
        print(f"   Fotos werden übersprungen (relativer Pfad wird als Text gespeichert).")
        print(f"   Lege den heruntergeladenen Google Drive-Ordner als '{FOTO_ORDNER}/' ab.\n")
    else:
        print(f"\n📁 Foto-Ordner: '{FOTO_ORDNER}' ✅")

    print(f"\n🔌 Verbinde mit Supabase...")
    db = create_client(SUPABASE_URL, SUPABASE_KEY)

    print(f"📂 Lade Excel: {EXCEL_FILE}")
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True)
    except FileNotFoundError:
        print(f"\n❌ Datei nicht gefunden: {EXCEL_FILE}")
        sys.exit(1)

    # ── BESTÄTIGUNG ───────────────────────────────────
    print("\n⚠️  ACHTUNG: Alle bestehenden Spieldaten werden gelöscht")
    print("   und neu aus der Excel-Datei importiert.")
    print("   Spieler-Accounts (Login) bleiben erhalten.")
    antwort = input("\n   Fortfahren? (j/n): ").strip().lower()
    if antwort != 'j':
        print("❌ Abgebrochen.")
        sys.exit(0)

    # ── SCHRITT 1: Alte Daten löschen ─────────────────
    print("\n🗑️  Schritt 1: Alte Daten löschen...")
    DUMMY = '00000000-0000-0000-0000-000000000000'
    for table in ['hand_statistik', 'transaktionen', 'spiel_teilnehmer', 'spiele']:
        db.table(table).delete().neq('id', DUMMY).execute()
        print(f"  🗑  {table} geleert")
    print("  ✅ Alte Daten gelöscht (Spieler-Accounts bleiben erhalten)")

    # ── SCHRITT 2: Spieler aktualisieren + Profilbilder ───
    print("\n📋 Schritt 2: Spieler aktualisieren (inkl. Profilbilder)...")
    res = db.table('spieler').select('id,name,email,ist_bank').execute()
    spieler_map = {}
    bank_id = None
    for s in res.data:
        spieler_map[s['name']] = s['id']
        if s['ist_bank']:
            bank_id = s['id']
    print(f"  ✅ {len(spieler_map)} Spieler in DB gefunden")

    ws_spieler = wb['Spieler']
    excel_spieler = [r for r in ws_spieler.iter_rows(values_only=True)
                     if r[0] and r[0] != 'Name']

    foto_konvertiert = foto_fehler = 0

    for row in excel_spieler:
        name, profilbild_roh, eintritt, email, aktiv = row

        # Profilbild konvertieren
        profilbild = None
        if profilbild_roh:
            print(f"  👤 {name}: Profilbild wird konvertiert...")
            profilbild = foto_zu_base64(profilbild_roh, PROFIL_MAX_PX)
            if profilbild and profilbild.startswith('data:'):
                foto_konvertiert += 1
            elif profilbild_roh:
                foto_fehler += 1

        updates = {}
        if profilbild is not None:  updates['profilbild']     = profilbild
        elif profilbild_roh:        updates['profilbild']     = str(profilbild_roh)
        if eintritt:                updates['eintrittsdatum'] = eintritt.date().isoformat()
        if email:                   updates['email']           = str(email)
        if aktiv is not None:       updates['aktiv']           = bool(aktiv)

        if name not in spieler_map:
            new = db.table('spieler').insert({
                'name':           name,
                'email':          str(email)      if email      else None,
                'profilbild':     profilbild or (str(profilbild_roh) if profilbild_roh else None),
                'eintrittsdatum': eintritt.date().isoformat() if eintritt else None,
                'aktiv':          bool(aktiv) if aktiv is not None else True,
                'ist_bank':       name == 'Bank',
            }).execute()
            spieler_map[name] = new.data[0]['id']
            if name == 'Bank':
                bank_id = new.data[0]['id']
            print(f"  ➕ Neuer Spieler: {name}")
        elif updates:
            db.table('spieler').update(updates).eq('id', spieler_map[name]).execute()

    print(f"  ✅ Spieler fertig – {foto_konvertiert} Profilbilder konvertiert"
          + (f", {foto_fehler} Fehler" if foto_fehler else ""))

    # ── SCHRITT 3: Spielverlauf ───────────────────────
    print("\n📋 Schritt 3: Spielverlauf importieren...")
    ws_verlauf = wb['Verlauf']
    verlauf_rows = [r for r in ws_verlauf.iter_rows(values_only=True)
                    if r[0] and r[0] != 'Key']

    abende = defaultdict(list)
    for row in verlauf_rows:
        key, datum, spieler, buyins, payout, f_pot, f_kasse, leihgabe, in_stat, sitz = row
        if datum and spieler:
            abende[datum.date()].append({
                'spieler':      spieler,
                'buyins':       int(buyins)     if buyins              else 0,
                'payout':       float(payout)   if payout              else 0.0,
                'f_pot':        float(f_pot)    if f_pot               else 5.0,
                'f_kasse':      float(f_kasse)  if f_kasse             else 2.0,
                'leihgabe':     float(leihgabe) if leihgabe            else 0.0,
                'in_statistik': bool(in_stat)   if in_stat is not None else True,
            })

    total = len(abende)
    spiele_neu = teilnehmer_neu = 0

    for i, (datum, teilnehmer) in enumerate(sorted(abende.items()), 1):
        if i % 20 == 0 or i == total:
            print(f"  ⏳ {i}/{total} Abende...")

        f_pot   = teilnehmer[0]['f_pot']
        f_kasse = teilnehmer[0]['f_kasse']

        spiel_res = db.table('spiele').insert({
            'datum':         datum.isoformat(),
            'abgeschlossen': True,
            'buyin_pot':     f_pot,
            'buyin_kassa':   f_kasse,
        }).execute()
        spiel_id = spiel_res.data[0]['id']
        spiele_neu += 1

        for t in teilnehmer:
            if t['spieler'] not in spieler_map:
                print(f"  ⚠️  Unbekannter Spieler: '{t['spieler']}' – übersprungen")
                continue
            db.table('spiel_teilnehmer').insert({
                'spiel_id':     spiel_id,
                'spieler_id':   spieler_map[t['spieler']],
                'buyins':       t['buyins'],
                'payout':       t['payout'],
                'leihgabe':     t['leihgabe'],
                'in_statistik': t['in_statistik'],
            }).execute()
            teilnehmer_neu += 1

    print(f"  ✅ {spiele_neu} Spielabende, {teilnehmer_neu} Teilnehmer-Einträge")

    # ── SCHRITT 4: Transaktionen ──────────────────────
    print("\n📋 Schritt 4: Transaktionen importieren...")
    ws_trans = wb['Transaktionen']
    trans_rows = [r for r in ws_trans.iter_rows(values_only=True)
                  if r[0] and r[0] != 'Key']

    trans_batch = []
    for row in trans_rows:
        key, datum, von, nach, betrag, kommentar = row
        if not datum or not betrag:
            continue
        trans_batch.append({
            'datum':           datum.date().isoformat(),
            'von_spieler_id':  spieler_map.get(str(von).strip())  if von  else None,
            'nach_spieler_id': spieler_map.get(str(nach).strip()) if nach else None,
            'betrag':          float(betrag),
            'kommentar':       str(kommentar) if kommentar else None,
        })

    for i in range(0, len(trans_batch), 50):
        db.table('transaktionen').insert(trans_batch[i:i+50]).execute()
    print(f"  ✅ {len(trans_batch)} Transaktionen")

    # ── SCHRITT 5: Hand Statistik + Beweisfotos ───────
    print("\n📋 Schritt 5: Hand Statistik importieren (inkl. Beweisfotos)...")
    ws_hands = wb['HandStatistik']
    hand_rows = [r for r in ws_hands.iter_rows(values_only=True)
                 if r[0] and r[0] != 'Key']

    foto_konv_h = foto_fehl_h = 0
    hand_batch = []

    for row in hand_rows:
        row = list(row) + [None] * 6
        key, datum, gewinner, hand, kommentar, beweisfoto_roh = row[:6]
        if not datum or not gewinner or not hand:
            continue

        gewinner_id = spieler_map.get(str(gewinner).strip())
        if not gewinner_id:
            print(f"  ⚠️  Unbekannter Gewinner: '{gewinner}' – übersprungen")
            continue

        # Beweisfoto konvertieren
        beweisfoto = None
        if beweisfoto_roh:
            beweisfoto = foto_zu_base64(beweisfoto_roh, BEWEIS_MAX_PX)
            if beweisfoto and beweisfoto.startswith('data:'):
                foto_konv_h += 1
            else:
                foto_fehl_h += 1

        hand_batch.append({
            'datum':       datum.date().isoformat(),
            'gewinner_id': gewinner_id,
            'hand':        str(hand),
            'kommentar':   str(kommentar) if kommentar else None,
            'beweisfoto':  beweisfoto,
        })

    for i in range(0, len(hand_batch), 50):
        db.table('hand_statistik').insert(hand_batch[i:i+50]).execute()
    print(f"  ✅ {len(hand_batch)} Hände"
          + (f" – {foto_konv_h} Beweisfotos konvertiert" if foto_konv_h else "")
          + (f", {foto_fehl_h} Fehler" if foto_fehl_h else ""))

    # ── ZUSAMMENFASSUNG ───────────────────────────────
    print("\n" + "=" * 50)
    print("🎉 Migration abgeschlossen!\n")
    print("Datenbank-Übersicht:")
    for table in ['spieler', 'spiele', 'spiel_teilnehmer', 'transaktionen', 'hand_statistik']:
        res = db.table(table).select('id', count='exact').execute()
        print(f"  {table:22s} {res.count:>4} Einträge")
    print("\n✅ App neu laden und Kontostände prüfen!")


if __name__ == '__main__':
    main()
